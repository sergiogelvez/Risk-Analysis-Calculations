import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl as xls

# load the excel workbook
# opportunity for improvement here
book = xls.load_workbook("libro.xlsx")
print(book)
# search sheet
print(book.sheetnames)
hoja = book.sheetnames[0]
#hoja = "hojita2"
#hoja = book.sheetnames[2]
# Code to select a sheet. 
# opportunity for improvement here
sheet = book[hoja]

# code to select square range. 
# opportunity for improvement here
cells = sheet["B2:AA27"]
#cells = sheet["A1:O15"]
#cells = sheet["A1:AH34"]


data_rows = []
for row in cells:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

B = np.array(data_rows)

rows, columns = B.shape

if rows != columns :
    print("Must be equal number of rows and columns")
    quit()

""" B = np.array([[0 , 1 , 2 , 0 , 2 , 0 , 1 , 0 , 0 , 0 , 0 , 3 , 1 , 1 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 0 , 0],\
[2 , 0 , 2 , 1 , 1 , 0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 3 , 1 , 1],\
[1 , 1 , 0 , 0 , 1 , 0 , 0 , 0 , 0 , 0 , 0 , 1 , 1 , 1 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 1 , 0],\
[0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 1 , 0 , 2 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 0],\
[1 , 0 , 1 , 0 , 0 , 0 , 0 , 0 , 0 , 1 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 0],\
[0 , 1 , 0 , 0 , 4 , 0 , 0 , 0 , 0 , 1 , 0 , 0 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 2 , 0 , 0],\
[1 , 1 , 1 , 2 , 0 , 0 , 0 , 1 , 3 , 2 , 1 , 1 , 3 , 3 , 3 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 2 , 3 , 1],\
[2 , 1 , 2 , 0 , 0 , 0 , 1 , 0 , 2 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 2 , 2],\
[0 , 1 , 1 , 2 , 0 , 0 , 0 , 0 , 0 , 1 , 0 , 0 , 3 , 2 , 4 , 3 , 3 , 3 , 3 , 3 , 3 , 3 , 3 , 1 , 4 , 2],\
[1 , 1 , 1 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0],\
[1 , 3 , 1 , 1 , 0 , 0 , 2 , 2 , 2 , 2 , 0 , 2 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1],\
[3 , 4 , 1 , 0 , 1 , 0 , 0 , 2 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 4 , 1 , 1 , 1 , 2 , 1 , 0],\
[1 , 1 , 1 , 0 , 2 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 3 , 1 , 1 , 1 , 1 , 1 , 3 , 0 , 1 , 0 , 0 , 0],\
[1 , 1 , 0 , 0 , 2 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 1 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 0 , 0],\
[1 , 1 , 0 , 0 , 2 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 2 , 2 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 0 , 1 , 1 , 1 , 1 , 1 , 0 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 1 , 1 , 1 , 1 , 0 , 0 , 0],\
[3 , 4 , 1 , 0 , 1 , 0 , 0 , 0 , 4 , 0 , 0 , 0 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 0 , 1 , 1 , 1 , 0 , 0 , 0],\
[0 , 0 , 0 , 2 , 1 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 1 , 1 , 0 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 1 , 1 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 1 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 0 , 0 , 0],\
[0 , 0 , 0 , 0 , 1 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 3 , 3 , 3 , 3 , 3 , 3 , 3 , 3 , 3 , 3 , 3 , 0 , 0 , 0],\
[1 , 2 , 0 , 2 , 1 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 1 , 0 , 0]])
 """

print(B)
rows, columns = B.shape 

# normalisation of direct influence matrix

print(rows, columns)
sums = np.zeros(rows)

for i in range(rows):
    sums[i] = sum(B[i, :])

max_sums_row = max(sums)
print(max_sums_row)

C = B / max_sums_row

print(C)
    
# C is the normalised matrix

I = np.identity(rows)
print(I) 

D = C @ np.linalg.inv(I - C)

print("\n\n")
print("Comprehensive influence matrix")
print(np.round(D, 2))
print("\n\n")
impact_degree = np.zeros(rows)
impact_risk_factor = np.zeros(columns)

# calculate impact dregree and impact risk for factors

for i in range(rows) :
    impact_degree[i] = sum(D[i, :])
    impact_risk_factor[i] = sum(D[:, i])

print("\n\n")
print("Impact degree")
print(impact_degree)
print("\n\n")
print("Impact risk factor")
print(impact_risk_factor)
print("\n\n")

# calculate centrality and causality

centrality = impact_degree + impact_risk_factor
causality = impact_degree - impact_risk_factor

print("\n\n")
print("Centrality")
print(centrality)
print("\n\n")
print("Causality")
print(causality)
print("\n\n")

# Causal diagram
main_plot = plt.plot(centrality, causality, "*r")
for i, val in enumerate(centrality) :
    plt.annotate(f"\u03B1{i + 1}", (centrality[i], causality[i]))
plt.show()

# Calculate overall influence matrix

H = np.identity(rows) + D
print(H)

# threslhold for reachability
l = 0.06
#l = 0.13
#l = 0.05

# arr[arr > threshold] = 0
print(H.shape)
K = np.empty_like(H, dtype=np.int32)

for i in range(rows):
    for j in range(columns):
        if H[i, j] >= l :
            K[i, j] = 1
        else :
            K[i, j] = 0

print("K")
print(K)


K_bak = np.copy(K)
#print(nonzero)
factores = rows 

structure = []
paso = 0
#Proceso de borrado
while factores > 0 :
    R = []
    for i in range(K.shape[0]):
        ri = []
        for j in range(K.shape[1]):
            if K[i, j] != 0 :
                ri.append(f"\u03B1{j + 1}")
        R.append(ri)

    S = []
    for i in range(K.shape[0]):
        si = []
        for j in range(K.shape[1]):
            if K[j, i] != 0 :
                si.append(f"\u03B1{j + 1}")
        S.append(si)

    print("\n\n")
    print("Reachable set:")
    for i in range(len(R)) :
        print(f"S{i + 1}: {R[i]}")

    print("\n\n")
    print("Antecedent set:")
    for i in range(len(S)) :
        print(f"S{i + 1}: {S[i]}")

    COLLSET = []
    for i in range(len(S)) :
        ci = list(set(S[i]) & set(R[i]))
        COLLSET.append(ci)

    print("\n\n")
    print("Collective set:")
    for i in range(len(COLLSET)) :
        print(f"S{i + 1}: {COLLSET[i]}")

    # compare R set with collective ser:
    print(f"\n\nPaso {paso}")
    borrar = []
    order = []
    for i in range(K.shape[0]) :
        R_set = set(R[i])
        Cl_set = set(COLLSET[i])
        if R_set == Cl_set and len(R[i]) > 0     :
            print(f"S{i + 1} is removed")
            borrar.append(i)
            order.append(f"S{i + 1}")
            factores -= 1
    structure.append(order)

    for i in range(len(borrar)):
        K[borrar[i], :] = 0
        K[:, borrar[i]] = 0

    paso += 1
    _ = input()

print("\n\n")
print("Hierarchy")
for level in structure:
    print(level)






'''
A = []
for i in range(len(R)) :
    R_set = set(R[i])
    Cl_set = set(COLLSET[i])
    if R_set == Cl_set :
        print(f"S{i} is removed")
    else :
        risk_dict = dict(name = f"S{i}", ri = R[i], si = S[i], cli = COLLSET[i])
        A.append(risk_dict)
'''